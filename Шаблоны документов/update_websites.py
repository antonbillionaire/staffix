import openpyxl
from openpyxl.styles import Font

verified = {
    "Vatan Tibbiyot Markazi": "https://vatantibbiyot.uz/",
    "Vitamed": "https://www.vitamed.uz/",
    "Nano Medical": "https://nanomedicalclinic.com/",
    "American dentistry": "https://americandentistry.uz/",
    "Smile boutique": "https://www.smileboutique.uz/",
    "Stomaservice": "https://stomaservice.uz/",
    "Dynasty": "https://dynasty.dental/",
    "X Dent": "https://www.xdent.uz/",
    "Sadaf": "https://sadafdental.uz/",
    "Smile concept": "https://smileconcept.uz/",
    "Kan Dental Clinic": "https://kan.uz/",
    "Nazar": "https://nazarmed.uz/",
    "Smalto": "https://www.smalto.uz/",
    "OCTA Clinic": "https://octaclinic.uz/",
    "VIP DENTAL SERVIS": "https://vipclinic.uz/",
    "Iftixor Dental Clinic": "https://iftixordental.uz/",
    "SERGO-DENTAL": "https://sakedental.uz/",
    "Dentcare": "https://www.instagram.com/dentcare.uz/",
    "Lanvin": "https://www.instagram.com/lanvin_clinic/",
    "BEGIZODA DENTAL CLINIC": "https://www.instagram.com/begizoda_dental_clinic/",
    "DENT STUDIO": "https://www.instagram.com/studio.dent_uz/",
    "DENTA-SERVICE": "https://www.denta-service.uz/",
    "DREAM SMILE DENTAL": "https://www.instagram.com/dsd.clinic/",
    "IRODA OILAVIY STOMATOLOGIYA": "https://www.instagram.com/iroda_stom/",
    "Al-Asnan-Dental": "https://alasnan.uz/",
    "Asmus Dental Clinic": "https://asmus.uz/",
    "Natural Hair Clinic": "https://www.instagram.com/natural_hair_clinic/",
    "DermaCenter": "https://dermacenter.uz/",
    "Diyor Medical Centre": "https://dmc.uz/",
    "Healthy Skin": "http://healthyskin.uz/ru/",
    "Heiberg Clinic": "https://www.instagram.com/heibergclinic_uz/",
    "My Cosmetology": "https://mycosmetology.uz/",
    "Planbaby clinic": "https://plan-baby.uz/",
    "Trionica": "https://trionica.uz/",
    "Vis Art": "https://www.visartmed.uz/",
    "FENIKS": "https://feniks.uz/",
    "Sinomed": "https://sino-med.uz/",
    "M beauty therapy clinic": "https://www.instagram.com/mbeautytherapy_uz/",
    "Inhype": "https://inhype-beauty-tashkent.uz/",
    "Opatra London": "https://www.instagram.com/opatra_london__tashkent/",
    "Prive7 Tashkent": "https://prive7.com/salons/prive7-beauty-express-tashkent/",
    "DERMATOLOGIYA VA KOSMETOLOGIYA": "https://www.facebook.com/DERMKOSMETOLOG/",
    "Bravo Medical Laboratory": "https://www.instagram.com/bravo.med.laboratory/",
    "Doctor Ashurov": "https://www.instagram.com/doctor_ashurovhospital/",
    "Medfit clinic": "https://www.instagram.com/medfit.uz/",
    "Nd Clinic": "https://ndcmedic.uz/",
    "Ibrat med clinic": "https://ibrat-med.uz/",
    "IYMON medical SBJ": "https://sbjmedical.uz/",
    "DMC-Klinika doktora Maksudovoy": "https://dmclinic.uz/",
    "International Shox Hospital": "https://shox.hospital/eng",
    "Horev": "https://horevmedical.uz/",
    "MDS": "https://mds.uz/",
    "Medion Clinic": "https://medion.uz/",
    "Orzu med Yunusobod": "https://orzumed.uz/",
    "Ihlos": "https://ihlos-doktor.uz/",
    "Orzu Medical Fotima Sulton": "https://orzumed.uz/",
    "Hippocrates Hospital": "https://www.instagram.com/hippocrates_hospital/",
    "Doktor D Hospital": "https://doctord.uz/",
    "ARK HOSPITAL": "https://ark.hospital/",
    "APOLLO HOSPITAL CLINIC": "https://apolloclinic.uz/",
    "CENTER OF DOCTOR BUBNOVSKY": "https://bubnovsky.org/tashkent/",
    "CITYMED CLINIC": "https://citymed.uz/",
    "DR. HASAN MEDICAL CENTER": "https://doctorhasan.uz/",
    "GENOTEXNOLOGIYA": "https://genotec.uz/",
    "IAC - International Allergy Center": "https://iac-tashkent.uz",
    "Dr. Steinke clinic": "https://steinkeivf.uz/",
    "EKO-IKSI": "https://eko-iksi.uz/",
    "Eramed EKO": "https://eramedeko.uz/",
    "Miracle IVF": "https://ecoivf.uz/",
    "Nanogen": "https://nanomedicalclinic.com/",
    "Asmo Clinic": "https://www.instagram.com/asmoclinic/",
    "Toshkent Tibbiy Diagnostika": "https://ttd.uz/",
    "Turk Med EKU MARKAZI": "https://www.instagram.com/turkmed_eku_markazi/",
    "Dr Mallaev Hospital": "https://www.instagram.com/dr.mallaevsclinic/",
    "BioLife": "https://biolife-clinic.uz/",
    "EviMedKids": "http://evimedkids.uz/",
    "Kidsmile": "https://kidsmile.uz/",
    "Soriz": "https://soriz-med.uz/",
    "Voice Lor clinic": "https://voicelor.uz/",
    "Lumen Vita": "https://lumenvita.uz/",
    "ERAmed": "https://eramed.uz/",
    "Dialab": "https://dialab.uz/",
    "Hi Tech Lab": "https://hitechlab.uz/",
    "Nazar Medical": "https://nazarmed.uz/",
    "Vedanta Medical": "https://vedanta.uz/",
    "CRYSTAL": "https://glaz.uz/",
    "Muqarnas Optic": "https://muqarnas-optic.uz/",
    "New Vision": "https://newvision.uz/",
    "Sihat Koz": "https://sihatkoz.uz/",
    "777 LOR": "https://lor777.uz/",
    "Yakkasaroy LOR servis 24/7": "https://www.instagram.com/yakkasaroy_lor_servis/",
    "Omega Eye": "https://www.instagram.com/omega_eye_clinic/",
    "Koz Med Servis": "https://www.instagram.com/kozmed.servis/",
    "Medanta": "https://medanta-clinic.uz/",
    "Koz nuri": "https://kuznuri.uz/",
    "Doktor Salimov": "https://allergy.uz/",
    "Otolor Hospital": "https://otolor.uz/",
    "Lor Plus Servis": "https://lorplus.uz/",
    "LOR PLUS SERVICE": "https://lorplus.uz/",
    "Lor Center": "https://toshkentlorcenter.uz/",
    "Profmed Lor": "https://profmed.uz/",
    "Hayat Hospital": "https://hayatmed.uz/",
    "Heart Team Clinic": "https://htclinic.uz/",
    "VITROS DIAGNOSTICS": "https://vitros.uz/",
    "FARHOD MADAD SHIFO": "https://t.me/farhodmadadshifoinfo",
    "Hasanov Abdurahmon": "https://www.instagram.com/khasanov_lor/",
    "Nigora Medical Star": "https://www.facebook.com/NigoraMedStar/",
    "Lor Klinika Dr Tobias": "https://www.instagram.com/dr.tobias_lor/",
    "Neyromed Servis": "https://www.instagram.com/neyromed_servis/",
    "Neyrohelp": "https://www.instagram.com/neyrohelp/",
    "Shaxzod medical": "https://shaxzod-medical.easyweek.io/",
}

print(f"Verified websites to add: {len(verified)}")

wb = openpyxl.load_workbook("Tashkent_ALL_Clinics_Obzvon.xlsx")
ws = wb.active

updated = 0
for r in range(2, ws.max_row + 1):
    name = ws.cell(row=r, column=3).value
    website = ws.cell(row=r, column=6).value

    if name and (not website or str(website).strip() == ""):
        name_clean = name.strip()
        name_lower = name_clean.lower().replace(" ", "").replace("-", "").replace(".", "")

        matched_url = None
        for vname, vurl in verified.items():
            vname_lower = vname.lower().replace(" ", "").replace("-", "").replace(".", "")
            if vname_lower == name_lower or vname_lower in name_lower or name_lower in vname_lower:
                matched_url = vurl
                break

        if matched_url:
            ws.cell(row=r, column=6, value=matched_url)
            ws.cell(row=r, column=6).font = Font(color="0563C1", underline="single")
            updated += 1

wb.save("Tashkent_ALL_Clinics_Obzvon.xlsx")

total = ws.max_row - 1
with_site = sum(1 for r in range(2, ws.max_row+1) if ws.cell(row=r, column=6).value and str(ws.cell(row=r, column=6).value).strip())
with_phone = sum(1 for r in range(2, ws.max_row+1) if ws.cell(row=r, column=4).value and str(ws.cell(row=r, column=4).value).strip() not in ["", "-", "None"])

print(f"Updated {updated} clinics with new websites")
print(f"Final: {total} clinics, {with_site} with website ({with_site*100//total}%), {with_phone} with phone ({with_phone*100//total}%)")
