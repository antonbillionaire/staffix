import openpyxl, re
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

BASE = "c:/Users/anton/Documents/GITHUB REPOSITORIUM"

records = []

# Source 1: clinics.uz (2880 lines)
with open(f"{BASE}/clinics_uz_all_categories.txt", "r", encoding="utf-8") as f:
    for line in f:
        line = line.strip()
        if not line or line.startswith("==="):
            continue
        parts = line.split("|")
        if len(parts) >= 3:
            cat, name = parts[0].strip(), parts[1].strip()
            phone = parts[2].strip() if len(parts) > 2 else ""
            addr = parts[3].strip() if len(parts) > 3 else ""
            web = parts[4].strip() if len(parts) > 4 else ""
            if name and name != "NAME":
                records.append((cat, name, phone, addr, web))

# Source 2: 2GIS
with open(f"{BASE}/tashkent_clinics_2gis.txt", "r", encoding="utf-8") as f:
    for line in f:
        line = line.strip()
        if not line or line.startswith("===") or line.startswith("CATEGORY"):
            continue
        parts = line.split("|")
        if len(parts) >= 2:
            cat, name = parts[0].strip(), parts[1].strip()
            phone = parts[2].strip() if len(parts) > 2 else ""
            addr = parts[3].strip() if len(parts) > 3 else ""
            web = parts[4].strip() if len(parts) > 4 else ""
            if name and name != "NAME":
                records.append((cat, name, phone, addr, web))

# Source 3: Yandex + extra
with open(f"{BASE}/tashkent_clinics_data.csv", "r", encoding="utf-8") as f:
    for line in f:
        line = line.strip()
        if not line or line.startswith("CATEGORY"):
            continue
        parts = line.split("|")
        if len(parts) >= 2:
            cat, name = parts[0].strip(), parts[1].strip()
            phone = parts[2].strip() if len(parts) > 2 else ""
            addr = parts[3].strip() if len(parts) > 3 else ""
            web = parts[4].strip() if len(parts) > 4 else ""
            if name:
                records.append((cat, name, phone, addr, web))

# Source 4: Excel with websites
wb_src = openpyxl.load_workbook(f"{BASE}/staffix/\u0428\u0430\u0431\u043b\u043e\u043d\u044b \u0434\u043e\u043a\u0443\u043c\u0435\u043d\u0442\u043e\u0432/tashkent_clinics_with_sites.xlsx")
ws_src = wb_src["tashkent_clinics_with_sites"]
excel_data = {}
for row in ws_src.iter_rows(min_row=2, max_row=ws_src.max_row, values_only=True):
    nm, ph, ad, ws_val = row
    if nm:
        key = re.sub(r"[^a-z0-9]", "", str(nm).strip().lower())
        excel_data[key] = {
            "phone": str(ph).strip().replace("\n"," ").replace("\r"," ") if ph else "",
            "website": str(ws_val).strip() if ws_val else ""
        }

print(f"Raw records loaded: {len(records)}")

# Normalize categories
cat_map = {}
for k in ["medical_centers", "medical centers"]:
    cat_map[k] = "Medical centers"
for k in ["gynecology", "гинекология"]:
    cat_map[k] = "Gynecology"
for k in ["ent_lor", "лор", "lor"]:
    cat_map[k] = "ENT (LOR)"
for k in ["neurology", "неврология"]:
    cat_map[k] = "Neurology"
for k in ["cardiology", "кардиология"]:
    cat_map[k] = "Cardiology"
for k in ["pediatrics", "педиатрия"]:
    cat_map[k] = "Pediatrics"
for k in ["urology", "урология"]:
    cat_map[k] = "Urology"
for k in ["ophthalmology", "офтальмология"]:
    cat_map[k] = "Ophthalmology"
for k in ["diagnostics", "diagnostic_center", "диагностика", "диагностический центр"]:
    cat_map[k] = "Diagnostics"
for k in ["cosmetology", "косметология"]:
    cat_map[k] = "Cosmetology"
for k in ["стоматология", "stomatology"]:
    cat_map[k] = "Stomatology"
for k in ["медицинские центры", "клиника (многопрофильная)"]:
    cat_map[k] = "Medical centers"

def norm_cat(c):
    return cat_map.get(c.lower().strip(), c.strip())

def norm_name(s):
    return re.sub(r"[^a-z0-9]", "", s.lower())

def phone_digits(p):
    if not p or p in ["", "-", "N/A", "None", "Not provided", "Not listed"]:
        return ""
    d = re.sub(r"[^0-9]", "", p.split(",")[0].split(";")[0])
    return d[-9:] if len(d) > 9 else d

# Dedup
clinic_db = {}
phone_map = {}

for cat_raw, name, phone, addr, web in records:
    cat = norm_cat(cat_raw)
    nk = norm_name(name)
    pd = phone_digits(phone)
    if phone in ["N/A", "Not provided", "Not listed", "None", "-"]:
        phone = ""

    if nk in clinic_db:
        ex = clinic_db[nk]
        if not ex["phone"] and phone: ex["phone"] = phone
        if not ex["address"] and addr: ex["address"] = addr
        if not ex["website"] and web: ex["website"] = web
        if cat and cat not in ex["categories"]: ex["categories"].append(cat)
    elif pd and pd in phone_map:
        other = phone_map[pd]
        if other in clinic_db:
            ex = clinic_db[other]
            if not ex["website"] and web: ex["website"] = web
            if not ex["address"] and addr: ex["address"] = addr
            if cat and cat not in ex["categories"]: ex["categories"].append(cat)
    else:
        clinic_db[nk] = {"name": name, "phone": phone, "address": addr, "website": web, "categories": [cat] if cat else []}
        if pd: phone_map[pd] = nk

# Enrich from Excel
for nk, data in clinic_db.items():
    if nk in excel_data:
        if not data["website"]: data["website"] = excel_data[nk]["website"]
        if not data["phone"]: data["phone"] = excel_data[nk]["phone"]

# Sort
cat_order = {"Stomatology":1, "Cosmetology":2, "Medical centers":3, "Gynecology":4, "Pediatrics":5,
             "Ophthalmology":6, "ENT (LOR)":7, "Cardiology":8, "Diagnostics":9, "Neurology":10, "Urology":11}

clinics = sorted(clinic_db.values(), key=lambda x: min(cat_order.get(c,99) for c in x["categories"]) if x["categories"] else 99)
print(f"Unique clinics: {len(clinics)}")

cat_counts = {}
for c in clinics:
    for cat in c["categories"]:
        cat_counts[cat] = cat_counts.get(cat, 0) + 1
for cat, cnt in sorted(cat_counts.items(), key=lambda x: cat_order.get(x[0], 99)):
    print(f"  {cat}: {cnt}")

with_phone = sum(1 for c in clinics if c["phone"])
with_site = sum(1 for c in clinics if c["website"])
print(f"With phone: {with_phone}, With website: {with_site}")

# Build Excel
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Clinics"

hf = Font(bold=True, size=11, color="FFFFFF")
hfill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
brd = Border(left=Side("thin"), right=Side("thin"), top=Side("thin"), bottom=Side("thin"))

cat_fills = {
    "Stomatology": PatternFill(start_color="E2EFDA", fill_type="solid"),
    "Cosmetology": PatternFill(start_color="FCE4D6", fill_type="solid"),
    "Medical centers": PatternFill(start_color="DDEBF7", fill_type="solid"),
    "Gynecology": PatternFill(start_color="F8CBAD", fill_type="solid"),
    "Pediatrics": PatternFill(start_color="D9E2F3", fill_type="solid"),
    "Ophthalmology": PatternFill(start_color="E2F0D9", fill_type="solid"),
    "ENT (LOR)": PatternFill(start_color="FFF2CC", fill_type="solid"),
    "Cardiology": PatternFill(start_color="D6DCE4", fill_type="solid"),
    "Diagnostics": PatternFill(start_color="EDEDED", fill_type="solid"),
    "Neurology": PatternFill(start_color="E2EFDA", fill_type="solid"),
    "Urology": PatternFill(start_color="DDEBF7", fill_type="solid"),
}

headers = ["No", "Category", "Name", "Phone", "Address", "Website", "Result", "Callback", "Contact Name", "Notes"]
for col, h in enumerate(headers, 1):
    c = ws.cell(row=1, column=col, value=h)
    c.font = hf; c.fill = hfill; c.border = brd
    c.alignment = Alignment(horizontal="center", vertical="center")

ws.column_dimensions["A"].width = 6
ws.column_dimensions["B"].width = 28
ws.column_dimensions["C"].width = 42
ws.column_dimensions["D"].width = 32
ws.column_dimensions["E"].width = 48
ws.column_dimensions["F"].width = 42
ws.column_dimensions["G"].width = 16
ws.column_dimensions["H"].width = 14
ws.column_dimensions["I"].width = 18
ws.column_dimensions["J"].width = 25

for i, c in enumerate(clinics, 1):
    r = i + 1
    primary = c["categories"][0] if c["categories"] else ""
    fill = cat_fills.get(primary, PatternFill(start_color="FFFFFF", fill_type="solid"))

    ws.cell(row=r, column=1, value=i).border = brd
    ws.cell(row=r, column=2, value=", ".join(c["categories"])).border = brd
    ws.cell(row=r, column=3, value=c["name"]).border = brd
    ws.cell(row=r, column=4, value=c["phone"]).border = brd
    ws.cell(row=r, column=5, value=c["address"]).border = brd
    c6 = ws.cell(row=r, column=6, value=c["website"])
    c6.border = brd
    if c["website"]:
        c6.font = Font(color="0563C1", underline="single")
    for col in range(7, 11):
        ws.cell(row=r, column=col, value="").border = brd
    for col in range(1, 11):
        ws.cell(row=r, column=col).fill = fill

ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:J{len(clinics)+1}"

dv = DataValidation(type="list", formula1='"Meeting,Callback,Refused,No Answer"', allow_blank=True)
ws.add_data_validation(dv)
for r in range(2, len(clinics)+2):
    dv.add(ws.cell(row=r, column=7))

outpath = f"{BASE}/staffix/\u0428\u0430\u0431\u043b\u043e\u043d\u044b \u0434\u043e\u043a\u0443\u043c\u0435\u043d\u0442\u043e\u0432/Tashkent_ALL_Clinics_Obzvon.xlsx"
wb.save(outpath)
print(f"\nSaved: {outpath}")
